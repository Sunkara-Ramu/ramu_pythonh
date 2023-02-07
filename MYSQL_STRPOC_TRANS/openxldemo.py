import openpyxl
wb=openpyxl.load_workbook("..//data/Book2.xlsx")
sheet=wb['Sheet2']
data=sheet['A3'].value
data3=sheet.cell(row=2,column=3).value
data1=sheet['A1:A10']
print(data)
print(data1)
print(data3)
#Access all cells in row 1
print(sheet.max_row)
print(sheet.max_column)
#for row in sheet.rows:
#print([data.value for data in row])
for i in range(2,12):
    print(sheet.cell(row=i,column=1).value)