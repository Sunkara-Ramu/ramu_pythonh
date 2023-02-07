from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
sheet.title="Hcl"
#sheet["A1"].value=10
#sheet["B2"].value="test"
#sheet.cell(row=3,column=3).value="Hcl data"
j=0
for i in range(10,60,10):
    j+=1
    sheet.cell(row=1,column=j).value=i

for row in sheet.iter.rows(min_row=1,max_row=5,max_col=2,min_col=2):
    for r in row:
        r.value=1
wb.save("..//data//demoopenxlwrite.xlsx")
